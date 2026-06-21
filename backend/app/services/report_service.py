"""
Report generation service for attendance data export.
"""

import io
import csv
from datetime import date, timedelta
from typing import List, Optional

from sqlalchemy.orm import Session
from sqlalchemy import func

from ..models.student import Student
from ..models.attendance import Attendance
from ..utils.timezone import get_vn_today


def get_attendance_summary(
    db: Session,
    start_date: date,
    end_date: date,
    class_name: Optional[str] = None,
) -> dict:
    """Generate attendance summary statistics for a date range."""
    query = (
        db.query(Attendance)
        .join(Student)
        .filter(func.date(Attendance.checkin_time) >= start_date, func.date(Attendance.checkin_time) <= end_date)
    )

    if class_name:
        query = query.filter(Student.class_name == class_name)

    records = query.all()

    total_records = len(records)
    on_time = sum(1 for r in records if r.status == "on_time")
    late = sum(1 for r in records if r.status == "late")
    early_leave = sum(1 for r in records if r.status == "early_leave")

    # Unique students who attended
    unique_students = len(set(r.student_id for r in records))

    # Total active students
    std_query = db.query(Student).filter(Student.is_active == True)
    if class_name:
        std_query = std_query.filter(Student.class_name == class_name)
    total_students = std_query.count()

    # Days in range
    days = (end_date - start_date).days + 1

    return {
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "total_records": total_records,
        "unique_employees": unique_students, # Keep this name to not break frontend immediately if they expect it
        "total_employees": total_students,
        "on_time": on_time,
        "late": late,
        "early_leave": early_leave,
        "days": days,
        "attendance_rate": round(unique_students / max(total_students, 1) * 100, 1),
    }


def generate_csv_report(
    db: Session,
    start_date: date,
    end_date: date,
    class_name: Optional[str] = None,
    student_id: Optional[int] = None,
) -> str:
    """Generate a CSV report of attendance records."""
    query = (
        db.query(Attendance)
        .join(Student)
        .filter(func.date(Attendance.checkin_time) >= start_date, func.date(Attendance.checkin_time) <= end_date)
        .order_by(func.date(Attendance.checkin_time), Student.full_name)
    )

    if class_name:
        query = query.filter(Student.class_name == class_name)
    if student_id:
        query = query.filter(Attendance.student_id == student_id)

    records = query.all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Ngày", "MSSV", "Họ tên", "Lớp",
        "Giờ điểm danh", "Trạng thái", "Độ tin cậy", "Ghi chú"
    ])

    for record in records:
        std = record.student
        writer.writerow([
            record.checkin_time.date().isoformat() if record.checkin_time else "",
            std.student_code,
            std.full_name,
            std.class_name or "",
            record.checkin_time.strftime("%H:%M:%S") if record.checkin_time else "",
            _status_label(record.status),
            f"{record.confidence:.1%}" if record.confidence else "",
            record.note or "",
        ])

    return output.getvalue()


def generate_excel_report(
    db: Session,
    start_date: date,
    end_date: date,
    class_name: Optional[str] = None,
    student_id: Optional[int] = None,
) -> bytes:
    """Generate an Excel report of attendance records."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    query = (
        db.query(Attendance)
        .join(Student)
        .filter(func.date(Attendance.checkin_time) >= start_date, func.date(Attendance.checkin_time) <= end_date)
        .order_by(func.date(Attendance.checkin_time), Student.full_name)
    )

    if class_name:
        query = query.filter(Student.class_name == class_name)
    if student_id:
        query = query.filter(Attendance.student_id == student_id)

    records = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo điểm danh"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"BÁO CÁO ĐIỂM DANH ({start_date.isoformat()} - {end_date.isoformat()})"
    title_cell.font = Font(name="Arial", size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Ngày", "MSSV", "Họ tên", "Lớp",
        "Giờ điểm danh", "Trạng thái", "Độ tin cậy", "Ghi chú"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    for row_idx, record in enumerate(records, 4):
        std = record.student
        row_data = [
            record.checkin_time.date().isoformat() if record.checkin_time else "",
            std.student_code,
            std.full_name,
            std.class_name or "",
            record.checkin_time.strftime("%H:%M:%S") if record.checkin_time else "",
            _status_label(record.status),
            f"{record.confidence:.1%}" if record.confidence else "",
            record.note or "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 3, 30)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def get_daily_trend(db: Session, days: int = 7, class_name: Optional[str] = None) -> list:
    """Get daily attendance count trend for the last N days."""
    today = get_vn_today()
    result = []

    for i in range(days - 1, -1, -1):
        d = today - timedelta(days=i)
        query = (
            db.query(func.count(Attendance.id))
            .join(Student)
            .filter(func.date(Attendance.checkin_time) == str(d))
        )
        if class_name:
            query = query.filter(Student.class_name == class_name)

        count = query.scalar() or 0
        result.append({"date": d.isoformat(), "count": count})

    return result


def get_department_stats(db: Session, target_date: date = None) -> list:
    """Get attendance breakdown by class for a specific date."""
    if target_date is None:
        target_date = get_vn_today()

    classes = (
        db.query(Student.class_name)
        .filter(Student.is_active == True, Student.class_name.isnot(None))
        .distinct()
        .all()
    )

    result = []
    for (cls,) in classes:
        total = db.query(Student).filter(
            Student.class_name == cls, Student.is_active == True
        ).count()
        attended = (
            db.query(func.count(Attendance.id))
            .join(Student)
            .filter(func.date(Attendance.checkin_time) == str(target_date), Student.class_name == cls)
            .scalar() or 0
        )
        result.append({
            "department": cls, # Keep key as department for frontend compatibility for now
            "total": total,
            "attended": attended,
            "rate": round(attended / max(total, 1) * 100, 1),
        })

    return result


def _status_label(status: str) -> str:
    labels = {
        "on_time": "Có mặt",
        "late": "Đi muộn",
        "absent": "Vắng mặt",
    }
    return labels.get(status, status)
